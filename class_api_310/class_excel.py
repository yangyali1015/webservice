import openpyxl
from openpyxl import load_workbook
class Excel():
    '''这是一个读取测试数据的存储表的类'''
    def __init__(self,filename,sheetname):
        self.filename=filename
        self.sheetname=sheetname
    def read(self):
        '''这是一个读取表单的函数
                sheet是表单名'''
        all_row_list = []  # 定义一个空列表用于装每行的列表
        wb = load_workbook(self.filename)  # 打开文件
        sheet_1 = wb[self.sheetname]  # 读取表单
        for i in range(2,sheet_1.max_row+1):  # i为行，设置循环的初始值为2
            row_list = []  # 此处定义一个每行的列表，用于存储每行的数据，每次循环完都清空
            for j in range(1, sheet_1.max_column + 1):  # 当i从1开始循环的时候，列数也开始从1循环
                cell = sheet_1.cell(i, j).value  # 每次循环都读取每行单元格的内容
                row_list.append(cell)  # 读取一行就加到行的列表里
            all_row_list.append(row_list)  # 每次循环都将行列表加到大列表里
        return all_row_list  # 返回大列表

    def write(self,row,column,content):
        wb = load_workbook(self.filename)
        sheet = wb[self.sheetname]
        value=sheet.cell(row,column,content)
        wb.save(self.filename)
        wb.close()
if __name__ == '__main__':
    print(Excel('case.xlsx','Sheet1').read())
    Excel('case.xlsx', 'Sheet1').write(2,8,8)