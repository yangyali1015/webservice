import unittest
import openpyxl
import logging
from configparser import ConfigParser
from openpyxl import workbook
from openpyxl import load_workbook
from ddt import ddt,data,unpack
# wb=workbook.Workbook()
# wb.save('练习.xlsx')
wb=load_workbook('练习.xlsx')
sheet=wb['Sheet']
# cell=sheet.cell(1,1,'caseid')
# cell=sheet.cell(1,2,'case')
# cell=sheet.cell(1,3,'a')
# cell=sheet.cell(1,4,'b')
# cell=sheet.cell(1,5,'expect')
# cell=sheet.cell(1,6,'result')
# cell=sheet.cell(2,1,1)
# cell=sheet.cell(3,1,2)
# cell=sheet.cell(4,1,3)
# wb.save('练习.xlsx')
# # wb.close()
# expect=sheet.cell(2,5).value
# print(expect)
mylogger=logging.getLogger('今天的日志')
mylogger.setLevel('INFO')
ch=logging.StreamHandler()
ch.setLevel("INFO")
fh=logging.FileHandler('自己瞎扯淡.log','w+')
fh.setLevel('INFO')
mylogger.addHandler(ch)
mylogger.addHandler(fh)
mylogger.removeHandler(fh)
mylogger.removeHandler(ch)
testcase_1=[[1,2,3],[2,3,5]]
@ddt
class Testcase(unittest.TestCase):
    def setUp(self):
        print('开始执行测试用例啦。。。。')
    def tearDown(self):
        print('。。。。用例执行完了啊。。。')
    @data(*testcase_1)
    @unpack
    def test_1(self,a,b,expect):
        c=a+b
        self.assertEqual(expect,c)




