import openpyxl
from openpyxl import load_workbook
from class_api_312 . conf_class import configer
from class_api_312 import path_class
class Doexcel:
    def __init__(self,filename,sheetname,sheetname_1):
        self.filename=filename
        self.sheetname=sheetname
        self.sheetname_1=sheetname_1
    def read(self):
              # xunhuan = 1
              # while xunhuan <= 10:
              #打开文件和表单
              wb = load_workbook(self.filename)
              sheet = wb[self.sheetname]
              sheet_1=wb[self.sheetname_1]
              #定义一个大列表存储数据
              finally_case=[]
              cf=configer('D:\\untitled\class_api\con_1.conf','妈呀','button')
              button = cf.get_other()
              if button=='all':
                for i in range(2,sheet.max_row+1):
                    case={}#每一行数据存在字典里面
                    case['case_id']=sheet.cell(i,1).value
                    case['Module'] = sheet.cell(i,2).value
                    case['URL'] = sheet.cell(i,3).value
                    case['Title'] = sheet.cell(i,4).value
                    case['Method'] = sheet.cell(i,5).value

                    if sheet.cell(i,6).value.find('tel')!=-1:#判断参数里面字符串是否含有tel
                        case['Params']=sheet.cell(i,6).value.replace('tel',str(sheet_1.cell(1,2).value))#如果含有就替换为第二个表单的手机号
                        sheet_1.cell(1,2,(sheet_1.cell(1,2).value)+1)
                    else:
                        case['Params'] = sheet.cell(i, 6).value
                    case['Sql'] = sheet.cell(i, 7).value
                    case['ExcepectedResult'] = sheet.cell(i, 8).value
                    finally_case.append(case)
              wb.save(self.filename)
              wb.close()

              return finally_case

    def write_back(self,i,j,content):#写入收据
        wb=load_workbook(self.filename)
        sheet=wb[self.sheetname]

        cell=sheet.cell(i,j,content)

        wb.save(self.filename)
        wb.close()

if __name__ == '__main__':
    c=Doexcel('case_2.xlsx','Sheet1','Sheet2')
    print(c.read())


