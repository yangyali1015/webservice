from openpyxl import load_workbook
import configparser
from configparser import ConfigParser
# 创建对象
cf=ConfigParser()
# 打开文件 read()
cf.read("lemon.conf",encoding="utf-8")
# 读取内容
button=cf.get("TestCase","button")
print(button)

# 读测试用例类
class DoExcel():
    def raad_excel(self,file_name,sheet_name):
        wb=load_workbook(file_name)     #打开工作薄
        sheet=wb[sheet_name]              #定位表单
        # 嵌套到字典中
        # test_data=[] #所有数据将要存在这个大字典中
        # for i in range(2,sheet.max_row+1):   #获取行数
        #    row_data = {}   #每一行数据都存在子字典中
        #    row_data['CaseID'] = sheet.cell(i,1).value
        #    row_data['Title'] =  sheet.cell(i, 2).value
        #    row_data['Module'] = sheet.cell(i, 3).value
        #    row_data['TestDate'] = sheet.cell(i, 4).value
        #    row_data['ExpectedResult'] = sheet.cell(i,5).value
        #    row_data['TActualResulte'] = sheet.cell(i, 6).value
        #    row_data['TestReuslt'] = sheet.cell(i, 7).value
        #    test_data.append(row_data)
        # return test_data

        # 嵌套到列表中
        # test_data=[]  #所有数据将要存在这个大列表中
        if button=='1':
           for j in range(2,sheet.max_row+1):   #获取行数
             row_data = []   #每一行数据都存在子列表中
           #   for i in range(1,sheet.max_column+1):  #获取列数
           #      row_data.append(sheet.cell(j,i).value)
           #   test_data.append(row_data)
           # return test_data
        else:
            test_data = []
            for i in eval(button): #如果button不等于1 eval（）之后就是一个列表
                  # 所有数据将要存在这个大列表中
                # for j in range(2,i):   #获取行数
                  row_data = []   #每一行数据都存在子列表中
                  for a in range(1,sheet.max_column+1):  #获取列数
                    row_data.append(sheet.cell(i+1,a).value)
            test_data.append(row_data)
            return test_data






if __name__=="__main__":
    p=DoExcel()
    res=p.raad_excel("接口测试用例.xlsx","TestCase")
    print(res)
