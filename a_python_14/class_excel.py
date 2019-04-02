import openpyxl
from openpyxl import load_workbook
from openpyxl import workbook
wb =workbook.Workbook()
wb.save('class_20190224.xlsx')
wb_1=load_workbook('class_20190224.xlsx')
wb.save('class_20190224.xlsx')
sheet=wb_1['Sheet']
cell=sheet.cell(2,3).value
print(cell)
