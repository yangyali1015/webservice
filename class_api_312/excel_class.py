import openpyxl
from openpyxl import load_workbook
from class_api_312 . conf_class import configer
from class_api_312 import path_class
class Doexcel:
    def __init__(self,filename,sheetname):
        self.filename=filename
        self.sheetname=sheetname
    def read(self):
        wb=load_workbook(self.filename)
        sheet=wb[self.sheetname]

        finally_case=[]
        cf=configer('D:\\untitled\class_接口作业_312\con_1.conf','妈呀','button')
        button = cf.get_other()
        for i in button:
            case=[]
            for j in range(1,sheet.max_column):
                cell=sheet.cell(i+1,j).value
                case.append(cell)
            finally_case.append(case)
        return finally_case

    def write_back(self,i,j,content):
        wb=load_workbook(self.filename)
        sheet=wb[self.sheetname]

        cell=sheet.cell(i,j,content)

        wb.save(self.filename)
        wb.close()

if __name__ == '__main__':
    a=Doexcel('case.xlsx','Sheet1')
    print(a.read())
    a.write_back(9,9,'hahah')