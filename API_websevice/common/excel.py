from openpyxl import load_workbook
from API_websevice.common import path
from API_websevice.common.getdata import Getdata
class Excel():
    '''读取表格用例'''
    def __init__(self,filename,sheet1_name,sheet2_name='tel'):
        self.filename=filename
        self.sheet1_name=sheet1_name
        self.sheet2_name = sheet2_name
    def read_excel(self):
        wb=load_workbook(self.filename)
        sheet=wb[self.sheet1_name]
        sheet_1=wb[self.sheet2_name]


        finally_case=[]
        for i in range(2,sheet.max_row+1):
            case={}
            case['case_id'] = sheet.cell(i, 1).value
            case['Module'] = sheet.cell(i, 2).value
            case['URL'] = sheet.cell(i, 3).value
            if sheet.cell(i, 3).value.find('url') != -1:
                case['URL'] = sheet.cell(i, 3).value.replace('url',str(sheet_1.cell(4, 2).value))  # 如果含有就替换为第二个表单的手机号
            case['Title'] = sheet.cell(i, 4).value
            case['Method'] = sheet.cell(i, 5).value
            case['Params'] = sheet.cell(i, 6).value
            if sheet.cell(i, 6).value.find('###') != -1:
                case['Params'] = sheet.cell(i, 6).value.replace('###',str(sheet_1.cell(1, 2).value))  # 如果含有就替换为第二个表单的手机号
                sheet_1.cell(1, 2, (sheet_1.cell(1, 2).value) + 1)
            elif sheet.cell(i, 6).value.find('user_uid') != -1:
                case['Params'] = sheet.cell(i, 6).value.replace('user_uid',str(sheet_1.cell(2, 2).value))  # 如果含有就替换为第二个表单的用户名
                sheet_1.cell(2, 2, (sheet_1.cell(2, 2).value) + 1)
            # elif sheet.cell(i, 6).value.find('ID_card') != -1:
            #     case['Params'] = sheet.cell(i, 6).value.replace('ID_card',str(sheet_1.cell(3, 2).value))  # 如果含有就替换为第二个表单的用户名
            #     sheet_1.cell(3, 2, str(int(sheet_1.cell(3, 2).value) + 10000))
            else:
                case['Params'] = sheet.cell(i, 6).value
            case['Sql'] = sheet.cell(i, 7).value
            case['ExcepectedResult'] = sheet.cell(i, 8).value
            finally_case.append(case)
            wb.save(self.filename)
            wb.close()
        return finally_case

    def write(self,row,column,con):
        wb=load_workbook(self.filename)
        sheet=wb[self.sheet1_name]
        sheet.cell(row,column,con)
        wb.save(self.filename)
        wb.close()

if __name__ == '__main__':
    a=Excel(path.excel_path,'bindBankCard','tel')
    print(a.read_excel())