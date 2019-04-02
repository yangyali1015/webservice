#第一题
import openpyxl
from openpyxl import load_workbook
from openpyxl import workbook
class Excel_Read_Write():
    '''这是一个可以对excel进行读写的类'''
    def __init__(self,filename,sheet):#
        '''filename是excel的文件名,sheet是表单名'''
        self.filename=filename
        self.sheet=sheet
    def read(self):
        '''这是一个读取表单的函数
        sheet是表单名'''
        all_row_list=[]#定义一个空列表用于装每行的列表
        wb=load_workbook(self.filename)#打开文件
        sheet_1=wb[self.sheet]#读取表单
        i=1#i为行，设置循环的初始值为1
        while i<=sheet_1.max_row:#只要行数小于最大行数都进入循环体
             row_list = []#此处定义一个每行的列表，用于存储每行的数据，每次循环完都清空
             for j in range(1,sheet_1.max_column+1):#当i从1开始循环的时候，列数也开始从1循环
                cell=sheet_1.cell(i,j).value#每次循环都读取每行单元格的内容
                row_list.append(cell)#读取一行就加到行的列表里
                all_row_list.append(row_list)#每次循环都将行列表加到大列表里
                i+=1#每次循环行数加1，一直到循环到最大行
        return all_row_list  #返回大列表
    def write(self,row,column,countent):
        '''在指定的单元格写入数据'''
        wb = load_workbook(self.filename)  # 打开文件
        #sheet_2=wb.get_sheet_by_name(self.filename)
        sheet_2=wb[self.sheet]#读取表单
        cell=sheet_2.cell(row,column,countent)#读取单元格并赋值
        wb.save(self.filename)#保存文件
    def create(self):
        wb=workbook.Workbook()#新建文件簿
        wb.save(self.filename)#命名并保存

if __name__ == '__main__':
    print(Excel_Read_Write('class_20190224.xlsx','Sheet').read())


#第二题
#单元格是字符串：读取值的类型是str
#单元格是整数：读取值的类型是int
#单元格是元祖：读取值的类型是str
#单元格是字典：读取值的类型是str
#单元格是浮点数：读取值的类型是float
#单元格是布尔值：读取值的类型是str
#单元格是列表：读取值的类型是str

#如果想变成原始类型就利用eval函数


