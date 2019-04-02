import openpyxl
from openpyxl import load_workbook
from configparser import  ConfigParser
from class_20190326 import  conf_20190326
from class_20190326.conf_20190326 import Config
cf = Config('con_20190326', 'CASE', 'case_id')
class DoExcel:
    '''这是一个读取表格的类'''
    def __init__(self,filename,sheetname):
        self.filename=filename
        self.sheetname=sheetname
    def read(self):
        wb=load_workbook(self.filename)
        sheet=wb[self.sheetname]
        sheet_1 = wb['tel']
        button=cf.getstr()
        cases_finally=[]
        if button=='all':
            for i in range(2,sheet.max_row+1):
                case={}
                case['case_id']=sheet.cell(i,1).value
                case['Module'] = sheet.cell(i, 2).value
                case['URL'] = sheet.cell(i, 3).value
                case['Title'] = sheet.cell(i, 4).value
                case['Method'] = sheet.cell(i, 5).value
                case['Params'] = sheet.cell(i, 6).value
                if sheet.cell(i,6).value.find('tel')!=-1:
                    case['Params'] = sheet.cell(i, 6).value.replace('tel',str(sheet_1.cell(1, 2).value))  # 如果含有就替换为第二个表单的手机号
                    sheet_1.cell(1, 2, (sheet_1.cell(1, 2).value) + 1)
                else:
                    case['Params'] = sheet.cell(i, 6).value
                case['Sql'] = sheet.cell(i, 7).value
                case['ExcepectedResult'] = sheet.cell(i, 8).value
                cases_finally.append(case)
                wb.save(self.filename)
                wb.close()
        return cases_finally

    def write(self,row,column,con):
        wb=load_workbook(self.filename)
        sheet=wb[self.sheetname]
        sheet.cell(row,column,con)
        wb.save(self.filename)
        wb.close()

if __name__ == '__main__':
    excel=DoExcel('Excelcase.xlsx','recharge')
    print(excel.read())